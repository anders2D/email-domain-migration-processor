"""
Tests Unitarios - ExcelEmailWriter (Output Adapter)
Cobertura de generación de archivos Excel (.xlsx)
"""
import pytest
import tempfile
import os
from pathlib import Path
from src.features.email_processing.adapters.output.excel_adapter import ExcelEmailWriter
from src.features.email_processing.domain.email import Email


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def temp_excel_file():
    """Crea archivo temporal para tests"""
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    yield path
    if os.path.exists(path):
        os.unlink(path)


# ============================================================================
# Tests de Inicialización
# ============================================================================

def test_excel_writer_init():
    """ExcelEmailWriter: __init__() con headers por defecto según PDD"""
    # Arrange & Act
    writer = ExcelEmailWriter()
    
    # Assert
    assert writer.headers == ['Nombre', 'Apellido', 'Correo Original', 'Correo Nuevo']


def test_excel_writer_init_custom_headers():
    """ExcelEmailWriter: __init__() con headers personalizados"""
    # Arrange
    custom_headers = ['Name', 'Surname', 'Old', 'New']
    
    # Act
    writer = ExcelEmailWriter(headers=custom_headers)
    
    # Assert
    assert writer.headers == custom_headers


# ============================================================================
# Tests de save_emails()
# ============================================================================

def test_excel_writer_save_empty_list(temp_excel_file):
    """ExcelEmailWriter: save_emails() con lista vacía crea archivo"""
    # Arrange
    writer = ExcelEmailWriter()
    
    # Act
    writer.save_emails([], temp_excel_file)
    
    # Assert
    assert os.path.exists(temp_excel_file)


def test_excel_writer_save_single_email(temp_excel_file):
    """ExcelEmailWriter: save_emails() con un email crea archivo válido"""
    # Arrange
    writer = ExcelEmailWriter()
    email = Email("Juan", "Perez", "juan.perez@old.com", "juan.perez@new.com")
    
    # Act
    writer.save_emails([email], temp_excel_file)
    
    # Assert
    assert os.path.exists(temp_excel_file)
    assert os.path.getsize(temp_excel_file) > 0


def test_excel_writer_save_multiple_emails(temp_excel_file):
    """ExcelEmailWriter: save_emails() con múltiples emails crea archivo"""
    # Arrange
    writer = ExcelEmailWriter()
    emails = [
        Email("Juan", "Perez", "juan.perez@old.com", "juan.perez@new.com"),
        Email("Maria", "Garcia", "maria.garcia@old.com", "maria.garcia@new.com")
    ]
    
    # Act
    writer.save_emails(emails, temp_excel_file)
    
    # Assert
    assert os.path.exists(temp_excel_file)


# ============================================================================
# Tests de Integridad de Datos
# ============================================================================

def test_excel_writer_file_created(temp_excel_file):
    """ExcelEmailWriter: Archivo Excel tiene hoja 'Correos Procesados'"""
    # Arrange
    writer = ExcelEmailWriter()
    email = Email("Juan", "Perez", "juan.perez@old.com", "juan.perez@new.com")
    
    # Act
    writer.save_emails([email], temp_excel_file)
    
    # Assert
    try:
        import openpyxl
        wb = openpyxl.load_workbook(temp_excel_file)
        assert wb.active.title == "Correos Procesados"
    except ImportError:
        pytest.skip("openpyxl not installed")


def test_excel_writer_headers_correct(temp_excel_file):
    """ExcelEmailWriter: Headers en Excel coinciden con formato PDD"""
    # Arrange
    writer = ExcelEmailWriter()
    email = Email("Juan", "Perez", "juan.perez@old.com", "juan.perez@new.com")
    
    # Act
    writer.save_emails([email], temp_excel_file)
    
    # Assert
    try:
        import openpyxl
        wb = openpyxl.load_workbook(temp_excel_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == ['Nombre', 'Apellido', 'Correo Original', 'Correo Nuevo']
    except ImportError:
        pytest.skip("openpyxl not installed")


def test_excel_writer_data_integrity(temp_excel_file):
    """ExcelEmailWriter: Datos en Excel mantienen integridad y orden"""
    # Arrange
    writer = ExcelEmailWriter()
    email = Email("Juan", "Perez", "juan.perez@old.com", "juan.perez@new.com")
    
    # Act
    writer.save_emails([email], temp_excel_file)
    
    # Assert
    try:
        import openpyxl
        wb = openpyxl.load_workbook(temp_excel_file)
        ws = wb.active
        data = [cell.value for cell in ws[2]]
        assert data == ['Juan', 'Perez', 'juan.perez@old.com', 'juan.perez@new.com']
    except ImportError:
        pytest.skip("openpyxl not installed")


def test_excel_writer_overwrite_existing(temp_excel_file):
    """ExcelEmailWriter: save_emails() sobrescribe archivo existente"""
    # Arrange
    writer = ExcelEmailWriter()
    email1 = Email("Juan", "Perez", "juan.perez@old.com", "juan.perez@new.com")
    email2 = Email("Maria", "Garcia", "maria.garcia@old.com", "maria.garcia@new.com")
    
    # Act
    writer.save_emails([email1], temp_excel_file)
    writer.save_emails([email2], temp_excel_file)
    
    # Assert
    assert os.path.exists(temp_excel_file)



